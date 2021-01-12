import sys
from openpyxl import load_workbook


def items_from_ranges(values):
    """Function to convert a range into a joined string of items

    Args:
        start (int): Start of range
        end (int): End of Range

    Returns:
        str: string of range items seperated by a comma
    """
    try:
        ranges = values.split(',')
    except ValueError:
        sys.exit(
            'Input not in correct format, must be a range seperate with a comma')

    formatted_ranges = []
    for item in ranges:
        if '-' in item:
            try:
                start, end = item.split('-')
            except ValueError:
                sys.exit(
                    'Input not in correct format, must be a range seperate with a hyphen')

            formatted_ranges.extend(
                [str(x) for x in list(range(int(start), int(end)+1))])
        else:
            formatted_ranges.append(str(item))
    # Convert Values
    return (',').join(formatted_ranges)


def format_workbook(path, col):
    wb = load_workbook(filename=path)
    skip_sheets = ['cover', 'vlans', 'vrfs']
    # Loop over sheets and format cols
    for sheet in wb.sheetnames:
        if sheet.lower() in skip_sheets:
            continue
        ws = wb[sheet]
        all_data = ws[col]
        for data in all_data:
            if data.value is not None:
                formatted = items_from_ranges(str(data.value))
                data.value = formatted

    wb.save('formatted.xlsx')
